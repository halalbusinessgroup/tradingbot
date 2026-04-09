"""Trades API — list, close, stats, XLSX export."""
import io
from datetime import datetime
from fastapi import APIRouter, Depends, Query, Response, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import desc
from typing import List, Optional
from app.database import get_db
from app.models.user import User
from app.models.trade import Trade
from app.models.strategy import Strategy
from app.models.api_key import ApiKey
from app.schemas.schemas import TradeOut
from app.core.deps import get_current_user

router = APIRouter(prefix="/api/trades", tags=["trades"])


@router.get("", response_model=List[TradeOut])
def list_trades(
    status: Optional[str] = Query(None),
    paper: Optional[bool] = Query(None),
    limit: int = 100,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(Trade).filter(Trade.user_id == user.id)
    if status:
        q = q.filter(Trade.status == status)
    if paper is not None:
        q = q.filter(Trade.paper_trade == paper)
    return q.order_by(desc(Trade.opened_at)).limit(limit).all()


@router.get("/stats")
def stats(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    open_count = db.query(Trade).filter(
        Trade.user_id == user.id, Trade.status == "OPEN", Trade.paper_trade == False
    ).count()
    closed = db.query(Trade).filter(
        Trade.user_id == user.id, Trade.status != "OPEN", Trade.paper_trade == False
    ).all()
    total_pnl = sum(t.pnl or 0 for t in closed)
    wins = sum(1 for t in closed if (t.pnl or 0) > 0)
    losses = sum(1 for t in closed if (t.pnl or 0) < 0)
    return {
        "open_trades": open_count,
        "closed_trades": len(closed),
        "total_pnl": round(total_pnl, 4),
        "wins": wins,
        "losses": losses,
        "win_rate": round((wins / len(closed)) * 100, 2) if closed else 0,
    }


@router.get("/strategy-stats")
def strategy_stats(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Per-strategy performance stats."""
    strategies = db.query(Strategy).filter(Strategy.user_id == user.id).all()
    result = []
    for s in strategies:
        closed = db.query(Trade).filter(
            Trade.strategy_id == s.id, Trade.status != "OPEN", Trade.paper_trade == False
        ).all()
        open_count = db.query(Trade).filter(
            Trade.strategy_id == s.id, Trade.status == "OPEN"
        ).count()
        total_pnl = sum(t.pnl or 0 for t in closed)
        wins = sum(1 for t in closed if (t.pnl or 0) > 0)
        durations = []
        for t in closed:
            if t.opened_at and t.closed_at:
                dur = (t.closed_at - t.opened_at).total_seconds() / 3600
                durations.append(dur)
        result.append({
            "id": s.id,
            "name": s.name,
            "is_active": s.is_active,
            "exchange": s.config.get("exchange", "binance"),
            "total_trades": len(closed),
            "open_trades": open_count,
            "wins": wins,
            "losses": len(closed) - wins,
            "win_rate": round((wins / len(closed)) * 100, 1) if closed else 0,
            "total_pnl": round(total_pnl, 4),
            "avg_duration_h": round(sum(durations) / len(durations), 1) if durations else 0,
        })
    return result


@router.post("/{trade_id}/close")
def manual_close(
    trade_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Manually close an open trade via market sell. Cancels TP/SL orders."""
    trade = db.query(Trade).filter(
        Trade.id == trade_id, Trade.user_id == user.id, Trade.status == "OPEN"
    ).first()
    if not trade:
        raise HTTPException(404, "Açıq trade tapılmadı")

    exit_price = None

    # Paper trade — just close it virtually
    if trade.paper_trade:
        # Use last known TP midpoint as virtual exit
        exit_price = trade.entry_price * 1.001  # slight positive, simulated
    else:
        # Real trade: execute market sell
        strat = db.query(Strategy).filter(Strategy.id == trade.strategy_id).first()
        exchange = strat.config.get("exchange", "binance") if strat else "binance"
        key = db.query(ApiKey).filter(
            ApiKey.user_id == user.id, ApiKey.exchange == exchange, ApiKey.is_active == True
        ).first()
        if not key:
            raise HTTPException(400, f"{exchange} API key tapılmadı")

        from app.core.security import decrypt_str
        from app.services.exchange_service import ExchangeService
        try:
            svc = ExchangeService(exchange, decrypt_str(key.api_key_enc), decrypt_str(key.api_secret_enc))
            # Cancel TP and SL orders first
            if trade.tp_order_id:
                svc.cancel_order(trade.symbol, trade.tp_order_id)
            if trade.sl_order_id:
                svc.cancel_order(trade.symbol, trade.sl_order_id)

            # Use actual free balance instead of stored qty to avoid NOTIONAL/LOT_SIZE errors
            # (fees may have reduced the exact amount slightly)
            exit_price = svc.get_price(trade.symbol)
            base_asset = trade.symbol.upper().replace("USDT", "").replace("BUSD", "").replace("BTC", "").rstrip()
            # Try to extract base from symbol more reliably
            for quote in ["USDT", "BUSD", "BTC", "ETH", "BNB"]:
                if trade.symbol.upper().endswith(quote):
                    base_asset = trade.symbol.upper()[:-len(quote)]
                    break

            sell_qty = trade.qty  # default
            try:
                bal = svc.fetch_balance()
                free_qty = bal.get(base_asset, {}).get("free", 0)
                if free_qty > 0:
                    sell_qty = free_qty
            except Exception:
                pass  # fallback to stored qty

            # Check minimum notional ($10 for most Binance pairs)
            MIN_NOTIONAL = 10.0
            if sell_qty * exit_price < MIN_NOTIONAL:
                # Position too small to sell — just close in DB
                pass
            else:
                svc.market_sell(trade.symbol, sell_qty)
        except Exception as e:
            err_str = str(e)
            if "-2015" in err_str or "Invalid API-key" in err_str:
                raise HTTPException(400, (
                    "Binance API xətası (-2015): API açarında 'Spot Trading' icazəsi aktiv deyil "
                    "və ya IP məhdudiyyəti var. Binance → API Management → Edit → "
                    "'Enable Spot & Margin Trading' seçin və IP məhdudiyyətini qaldırın."
                ))
            raise HTTPException(500, f"Trade bağlama xətası: {e}")

    # Update trade record
    trade.exit_price = exit_price
    trade.status = "CLOSED_MANUAL"
    trade.closed_at = datetime.utcnow()
    if exit_price:
        trade.pnl = round((exit_price - trade.entry_price) * trade.qty, 6)
        trade.pnl_percent = round(
            ((exit_price - trade.entry_price) / trade.entry_price) * 100, 4
        )
    db.commit()

    # Telegram notification
    from app.models.log import Log
    from app.services import telegram_service as tg
    db.add(Log(user_id=user.id, level="INFO",
               message=f"🔒 Manuel bağlama: {trade.symbol} @ {exit_price} | PnL={trade.pnl:+.4f} USDT"))
    db.commit()
    tg.send_message(user.telegram_chat_id, f"🔒 Trade bağlandı (Manuel)\n{trade.symbol} @ {exit_price}\nPnL: {trade.pnl:+.4f} USDT")

    return {"ok": True, "pnl": trade.pnl, "exit_price": exit_price}


@router.get("/export")
def export_trades(
    fmt: str = Query("xlsx", description="Format: xlsx or csv"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export all trade history as XLSX or CSV."""
    trades = db.query(Trade).filter(
        Trade.user_id == user.id, Trade.status != "OPEN"
    ).order_by(desc(Trade.opened_at)).all()

    rows = []
    for t in trades:
        rows.append({
            "ID": t.id,
            "Symbol": t.symbol,
            "Side": t.side,
            "Qty": t.qty,
            "Entry Price": t.entry_price,
            "Exit Price": t.exit_price,
            "TP Price": t.tp_price,
            "SL Price": t.sl_price,
            "Status": t.status,
            "PnL (USDT)": t.pnl,
            "PnL %": t.pnl_percent,
            "Paper Trade": "Yes" if t.paper_trade else "No",
            "Opened At": t.opened_at.strftime("%Y-%m-%d %H:%M:%S") if t.opened_at else "",
            "Closed At": t.closed_at.strftime("%Y-%m-%d %H:%M:%S") if t.closed_at else "",
        })

    if fmt == "csv":
        import csv
        buf = io.StringIO()
        if rows:
            writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel UTF-8
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=trades.csv"},
        )
    else:
        # XLSX
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"

        headers = list(rows[0].keys()) if rows else []
        header_fill = PatternFill("solid", fgColor="1a2a1a")
        header_font = Font(bold=True, color="22c55e")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(str(header)) + 4, 12)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, (key, val) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                # Color PnL column
                if key == "PnL (USDT)" and isinstance(val, (int, float)):
                    cell.font = Font(color="22c55e" if val >= 0 else "ef4444")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=trades.xlsx"},
        )
